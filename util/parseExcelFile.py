"""
------------------------------------
@Time : 2019/4/22 16:12
@Auth : linux超
@File : parseExcelFile.py
@IDE  : PyCharm
@Motto: Real warriors,dare to face the bleak warning,dare to face the incisive error!
------------------------------------
"""
from openpyxl import load_workbook
from config.conf import DATA_Path


class ParseExcel(object):

    def __init__(self):
        self.wk = load_workbook(DATA_Path)
        self.excelFile = DATA_Path

    def get_sheet_by_name(self, sheet_name):
        """获取sheet对象"""
        sheet = self.wk[sheet_name]
        return sheet

    def get_row_num(self, sheet):
        """获取有效数据的最大行号"""
        return sheet.max_row

    def get_cols_num(self, sheet):
        """获取有效数据的最大列号"""
        return sheet.max_column

    def get_row_values(self, sheet, row_num):
        """获取某一行的数据"""
        max_cols_num = self.get_cols_num(sheet)
        row_values = []
        for colsNum in range(1, max_cols_num + 1):
            value = sheet.cell(row_num, colsNum).value
            if value is None:
                value = ''
            row_values.append(value)
        return tuple(row_values)

    def get_column_values(self, sheet, column_num):
        """获取某一列的数据"""
        max_row_num = self.get_row_num(sheet)
        column_values = []
        for rowNum in range(2, max_row_num + 1):
            value = sheet.cell(rowNum, column_num).value
            if value is None:
                value = ''
            column_values.append(value)
        return tuple(column_values)

    def get_value_of_cell(self, sheet, row_num, column_num):
        """获取某一个单元格的数据"""
        value = sheet.cell(row_num, column_num).value
        if value is None:
            value = ''
        return value

    def get_all_values_of_sheet(self, sheet):
        """获取某一个sheet页的所有测试数据，返回一个元祖组成的列表"""
        max_row_num = self.get_row_num(sheet)
        column_num = self.get_cols_num(sheet)
        all_values = []
        for row in range(2, max_row_num + 1):
            row_values = []
            for column in range(1, column_num + 1):
                value = sheet.cell(row, column).value
                if value is None:
                    value = ''
                row_values.append(value)
            all_values.append(tuple(row_values))
        return all_values


if __name__ == '__main__':
    pass
